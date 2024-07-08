import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook
import os


scopes = ["https://www.googleapis.com/auth/spreadsheets" , "https://www.googleapis.com/auth/drive"]
creds = Credentials.from_service_account_file("credentials.json", scopes=scopes)
client = gspread.authorize(creds)

spreadsheet = client.open('Farm5. Айди квестов')
worksheet = spreadsheet.worksheet('Farm5')


def get_column_data(worksheet, column):
    """Get all data from a specific column."""
    return worksheet.col_values(column)

def update_columns(worksheet, initial_master_data, current_master_data, attached_columns):
    """
    Rearrange attached columns according to the rearranged master column.
    """
    # Create a map of initial master data to their indices
    initial_map = {value: idx for idx, value in enumerate(initial_master_data)}
    
    # Debug print
    print(f"Initial Map: {initial_map}")

    # Sort the attached columns based on the new order of master data
    for col in attached_columns:
        col_data = worksheet.col_values(col)
        
        # Ensure col_data is at least as long as current_master_data
        if len(col_data) < len(current_master_data):
            col_data.extend([''] * (len(current_master_data) - len(col_data)))
        
        sorted_col_data = [None] * len(current_master_data)
        for idx, value in enumerate(current_master_data):
            if value in initial_map:
                sorted_col_data[idx] = col_data[initial_map[value]]
            else:
                sorted_col_data[idx] = ""  # Handle missing values
        
        # Debug print
        print(f"Updated Column {col}: {sorted_col_data}")

        # Update the column with sorted data
        cell_list = worksheet.range(1, col, len(sorted_col_data), col)
        for cell, value in zip(cell_list, sorted_col_data):
            cell.value = value
        worksheet.update_cells(cell_list)

# Define the Master column number (1-based index)
master_column_number = 2

# Load the initial state from the Excel file
wb = load_workbook("master_column_initial_state.xlsx")
ws = wb["InitialState"]
initial_master_column_data = [cell.value for cell in ws['A']]

# Get the current state of the master column
current_master_column_data = get_column_data(worksheet, master_column_number)

# Define the attached columns numbers (1-based index)
attached_columns_indices = [7, 8, 9, 12, 13, 17, 18, 19, 24, 25]  # Example: Attach the second and third columns

# Call the function to update the attached columns
update_columns(worksheet, initial_master_column_data, current_master_column_data, attached_columns_indices)

file_path = "master_column_initial_state.xlsx"

try:
    os.remove(file_path)
    print(f"{file_path} has been removed successfully")
except FileNotFoundError:
    print(f"{file_path} does not exist")
except Exception as e:
    print(f"Error: {e}")